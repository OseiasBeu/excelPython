import os
# import curses
# window = curses.initscr()
# window.clear()
# window.refresh()
os.system('clear' if os.name == 'nt' else 'cls')

from openpyxl import Workbook
arquivo_excel = Workbook()


planilha1 = arquivo_excel.active

# Mudando o título de uma planilha
planilha1.title = "Gastos"

# Criando uma nova planilha
planilha2 = arquivo_excel.create_sheet("Ganhos",0)

# Exibe o nomes das planilhas contidas dentro do arquivo
print(arquivo_excel.sheetnames)

# Atribuindo valores a campos específicos:
planilha1['A1']= 'Categoria'
planilha1['B1']= 'Valor'
planilha1['A2']= 'Restaurante'
planilha1['B2']= 45.99

# Adicionando um grupo de valores a última linha escrita:

valores=[
    ("Categoria","Valor"),
    ("Restaurante", 45,99),
    ("Transporte", 208.45),
    ("Viagem", 558.54)
]

for linha in valores:
    planilha1.append(linha)


# Usando o método Cell para modificar os valores:
planilha1.cell(row=3,column=1,value=34.99)

# Adicionando fórmulas:
planilha1['C1']= '=SOMA(23;5) \n'

# LENDO DADOS DA PLANILHA:

#Utilizando em índices o nome das células como em um dicionário:
c1 = planilha1['C1'].value
print("valor de C1:" + c1)

#Utilizando o método cell
a1 = planilha1.cell(column=1, row=1)
print("Valor de A1:" + a1.value)

# TEMBÉM PODEMOS USAR UMA ESTRUTURA DE REPETIÇÃO PARA LER LINHA A LINHA:
max_linha = planilha1.max_row
print(max_linha)
max_coluna = planilha1.max_column
print(max_coluna)

for i in range(1,max_coluna +1):
    for j in range(1,max_coluna +1):
        print(planilha1.cell(row=i, column=j).value, end="-")


# SALVANDO A PLANILHA:
print(' \n SALVANDO ARQUIVO........ \n')
arquivo_excel.save('relatorio1.xlsx')
print('\n ARQUIVO: relatorio.xlsx - FOI SALVO COM SUCESSO!.......')
